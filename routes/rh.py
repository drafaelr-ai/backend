"""Blueprint do módulo Pessoal / RH — funcionários, convenções (CCT),
pagamentos de salário e encargos. Todas as rotas exigem JWT.
"""
import io
import csv
import json
import logging
from datetime import datetime, date

from flask import Blueprint, jsonify, request
from flask_jwt_extended import jwt_required, verify_jwt_in_request
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import joinedload

from extensions import db
from models.categoria_mo import CategoriaMO
from models.convencao_coletiva import ConvencaoColetiva
from models.convencao_valor import ConvencaoValor
from models.funcionario import Funcionario
from models.pagamento_salario import PagamentoSalario
from models.encargo import Encargo
from models.obra import Obra
from services import cct_parser_service, rh_service, storage_service
from services import get_current_user, user_has_access_to_obra, user_tem_modulo

logger = logging.getLogger(__name__)

rh_bp = Blueprint('rh', __name__, url_prefix='/rh')


@rh_bp.before_request
def _gate_modulo_rh():
    """Acesso ao módulo RH exige o módulo liberado (master sempre passa)."""
    if request.method == 'OPTIONS':
        return None
    verify_jwt_in_request()
    if not user_tem_modulo(get_current_user(), 'rh'):
        return jsonify({"erro": "Acesso negado: você não tem permissão para o módulo RH."}), 403

_ENCARGO_TIPOS = {'fgts', 'inss_darf', 'esocial_dae', 'outro'}
_PAG_TIPOS = {'salario', 'vale', 'outro'}


# ---------------------------------------------------------------- helpers

def _parse_date(valor):
    """Aceita 'YYYY-MM-DD' (ou ISO com hora) → date; None se vazio/ inválido."""
    if not valor:
        return None
    if isinstance(valor, date):
        return valor
    try:
        return datetime.fromisoformat(str(valor)[:10]).date()
    except Exception:
        return None


def _to_num(valor):
    """Converte número ou string BR ('2.640,00') em float. None se vazio."""
    if valor is None or valor == '':
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace('R$', '').strip()
    if ',' in s and '.' in s:          # 2.640,00 → 2640.00
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:                     # 2640,00 → 2640.00
        s = s.replace(',', '.')
    try:
        return float(s)
    except Exception:
        return None


def _parse_int_arg(nome, valor):
    """int(valor) com erro amigável (400) em vez de deixar ValueError estourar
    pro except genérico da rota (que devolveria 500). Retorna (valor, erro);
    `erro` é a tupla (response, status) pronta pra `return` quando inválido."""
    try:
        return int(valor), None
    except (TypeError, ValueError):
        return None, (jsonify({"erro": f"{nome} inválido"}), 400)


def _obra_ids_permitidos(user):
    """None = sem restrição (master/administrador, vê tudo). Lista = só essas
    obras (usuário comum); lista vazia = nenhuma obra liberada."""
    if user and user.role in ('master', 'administrador'):
        return None
    return [o.id for o in user.obras_permitidas] if user else []


def _restringir_por_obra(query, coluna_obra_id, user):
    """Restringe a query às obras permitidas do usuário (no-op p/ master/administrador)."""
    obra_ids = _obra_ids_permitidos(user)
    if obra_ids is None:
        return query
    return query.filter(coluna_obra_id.in_(obra_ids))


# ---------------------------------------------------------------- categorias

@rh_bp.route('/categorias', methods=['GET'])
@jwt_required()
def listar_categorias():
    try:
        cats = CategoriaMO.query.order_by(CategoriaMO.nome).all()
        return jsonify([c.to_dict() for c in cats]), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/categorias")
        return jsonify({"erro": "Erro ao listar categorias", "detalhe": str(e)}), 500


@rh_bp.route('/categorias', methods=['POST'])
@jwt_required()
def criar_categoria():
    """Criação manual de categoria (antes só nascia pelo wizard de CCT)."""
    try:
        dados = request.get_json(silent=True) or {}
        nome = (dados.get('nome') or '').strip()
        if not nome:
            return jsonify({"erro": "nome é obrigatório"}), 400
        cat = _resolver_categoria(nome)
        if dados.get('descricao') and not cat.descricao:
            cat.descricao = dados['descricao'].strip()
        db.session.commit()
        return jsonify(cat.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em POST /rh/categorias")
        return jsonify({"erro": "Erro ao criar categoria", "detalhe": str(e)}), 500


# ---------------------------------------------------------------- convenções

@rh_bp.route('/convencoes', methods=['GET'])
@jwt_required()
def listar_convencoes():
    try:
        convs = ConvencaoColetiva.query.order_by(ConvencaoColetiva.data_upload.desc()).all()
        contagens = dict(
            db.session.query(ConvencaoValor.convencao_id, db.func.count(ConvencaoValor.id))
            .group_by(ConvencaoValor.convencao_id)
            .all()
        )
        out = []
        for c in convs:
            d = c.to_dict()
            d['categorias_count'] = contagens.get(c.id, 0)
            out.append(d)
        return jsonify(out), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/convencoes")
        return jsonify({"erro": "Erro ao listar convenções", "detalhe": str(e)}), 500


@rh_bp.route('/convencoes/extrair', methods=['POST'])
@jwt_required()
def extrair_convencao():
    """Upload do PDF → roda o parser → retorna JSON de categorias. NÃO salva.

    Degradação graciosa: qualquer falha do parser (sem chave, timeout, resposta
    não-JSON) retorna 200 com categorias vazias + aviso, para o fluxo de revisão
    abrir em modo manual. Nunca 500/503 (e nunca 422, que o fetchWithAuth trata
    como sessão expirada e desloga o operador)."""
    _AVISO = "Não consegui ler a convenção automaticamente. Preencha as categorias manualmente."
    try:
        arquivo = request.files.get('arquivo') or request.files.get('file')
        if not arquivo:
            return jsonify({"erro": "arquivo (PDF) é obrigatório"}), 400
        resultado = cct_parser_service.parse_cct(arquivo)
        return jsonify(resultado), 200
    except Exception:
        logger.exception("Falha no parser de CCT — degradando p/ preenchimento manual")
        return jsonify({"categorias": [], "aviso": _AVISO}), 200


def _resolver_categoria(nome):
    """Encontra CategoriaMO por nome (case-insensitive) ou cria uma nova.

    Usa um SAVEPOINT (begin_nested) ao criar: se outra transação concorrente
    criou a mesma categoria (nome case-insensitive) entre o SELECT e o
    INSERT, o índice único de categoria_mo.lower(nome) dispara IntegrityError
    aqui — nesse caso só desfazemos o savepoint e reconsultamos a categoria
    já criada pela outra transação, em vez de duplicar."""
    nome = (nome or '').strip()
    if not nome:
        return None
    cat = CategoriaMO.query.filter(db.func.lower(CategoriaMO.nome) == nome.lower()).first()
    if cat:
        return cat
    try:
        with db.session.begin_nested():
            cat = CategoriaMO(nome=nome)
            db.session.add(cat)
            db.session.flush()
    except IntegrityError:
        cat = CategoriaMO.query.filter(db.func.lower(CategoriaMO.nome) == nome.lower()).first()
    return cat


@rh_bp.route('/convencoes', methods=['POST'])
@jwt_required()
def criar_convencao():
    """Salva a CCT confirmada: cria categorias faltantes, valores e benefícios,
    sobe o PDF ao Storage. Aceita multipart (com arquivo) ou JSON."""
    try:
        if request.content_type and 'multipart' in request.content_type:
            dados = request.form
            categorias = json.loads(dados.get('categorias') or '[]')
            arquivo = request.files.get('arquivo') or request.files.get('file')
        else:
            dados = request.get_json() or {}
            categorias = dados.get('categorias') or []
            arquivo = None

        uf = (dados.get('uf') or '').strip().upper()
        if not uf:
            return jsonify({"erro": "uf é obrigatório"}), 400
        vig_ini = _parse_date(dados.get('vigencia_inicio'))
        vig_fim = _parse_date(dados.get('vigencia_fim'))
        if not vig_ini or not vig_fim:
            return jsonify({"erro": "vigencia_inicio e vigencia_fim são obrigatórias (YYYY-MM-DD)"}), 400
        if vig_ini > vig_fim:
            return jsonify({"erro": "vigencia_inicio deve ser <= vigencia_fim"}), 400

        arquivo_url = None
        anexo_falhou = False
        if arquivo:
            try:
                arquivo_url = storage_service.upload_arquivo(arquivo, 'convencoes')
            except Exception as e:
                anexo_falhou = True
                logger.exception("Upload da CCT falhou (segue sem arquivo): %s", e)

        conv = ConvencaoColetiva(
            uf=uf,
            sindicato=(dados.get('sindicato') or '').strip() or None,
            vigencia_inicio=vig_ini,
            vigencia_fim=vig_fim,
            arquivo_url=arquivo_url,
            status='confirmada',
        )
        db.session.add(conv)
        db.session.flush()

        for cat in categorias:
            categoria = _resolver_categoria(cat.get('nome'))
            if not categoria:
                continue
            piso = _to_num(cat.get('piso_salarial'))
            if piso is None:
                continue
            db.session.add(ConvencaoValor(
                convencao_id=conv.id,
                categoria_id=categoria.id,
                piso_salarial=piso,
                beneficios=cat.get('beneficios') or [],
            ))

        db.session.commit()
        resp = conv.to_dict()
        if anexo_falhou:
            resp['aviso'] = "Convenção salva, mas o anexo (PDF) não pôde ser enviado."
        return jsonify(resp), 201
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em POST /rh/convencoes")
        return jsonify({"erro": "Erro ao salvar convenção", "detalhe": str(e)}), 500


@rh_bp.route('/convencoes/<int:conv_id>', methods=['GET'])
@jwt_required()
def obter_convencao(conv_id):
    try:
        conv = db.session.get(ConvencaoColetiva, conv_id)
        if not conv:
            return jsonify({"erro": "Convenção não encontrada"}), 404
        d = conv.to_dict()
        valores = ConvencaoValor.query.filter_by(convencao_id=conv.id).all()
        d['valores'] = [v.to_dict() for v in valores]
        return jsonify(d), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/convencoes/<id>")
        return jsonify({"erro": "Erro ao obter convenção", "detalhe": str(e)}), 500


@rh_bp.route('/convencoes/<int:conv_id>', methods=['PUT'])
@jwt_required()
def editar_convencao(conv_id):
    try:
        conv = db.session.get(ConvencaoColetiva, conv_id)
        if not conv:
            return jsonify({"erro": "Convenção não encontrada"}), 404
        dados = request.get_json() or {}

        if 'sindicato' in dados:
            conv.sindicato = (dados.get('sindicato') or '').strip() or None
        if 'uf' in dados and dados.get('uf'):
            conv.uf = dados['uf'].strip().upper()
        if dados.get('vigencia_inicio'):
            conv.vigencia_inicio = _parse_date(dados['vigencia_inicio']) or conv.vigencia_inicio
        if dados.get('vigencia_fim'):
            conv.vigencia_fim = _parse_date(dados['vigencia_fim']) or conv.vigencia_fim
        if conv.vigencia_inicio and conv.vigencia_fim and conv.vigencia_inicio > conv.vigencia_fim:
            return jsonify({"erro": "vigencia_inicio deve ser <= vigencia_fim"}), 400
        if dados.get('status') in ('rascunho', 'confirmada'):
            conv.status = dados['status']

        # Substitui os valores/benefícios se enviados
        if 'categorias' in dados or 'valores' in dados:
            entrada = dados.get('categorias') or dados.get('valores') or []
            ConvencaoValor.query.filter_by(convencao_id=conv.id).delete()
            for cat in entrada:
                categoria = _resolver_categoria(cat.get('nome') or cat.get('categoria_nome'))
                piso = _to_num(cat.get('piso_salarial'))
                if not categoria or piso is None:
                    continue
                db.session.add(ConvencaoValor(
                    convencao_id=conv.id,
                    categoria_id=categoria.id,
                    piso_salarial=piso,
                    beneficios=cat.get('beneficios') or [],
                ))

        db.session.commit()
        return jsonify(conv.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em PUT /rh/convencoes/<id>")
        return jsonify({"erro": "Erro ao editar convenção", "detalhe": str(e)}), 500


@rh_bp.route('/convencoes/<int:conv_id>', methods=['DELETE'])
@jwt_required()
def remover_convencao(conv_id):
    try:
        conv = db.session.get(ConvencaoColetiva, conv_id)
        if not conv:
            return jsonify({"erro": "Convenção não encontrada"}), 404
        db.session.delete(conv)   # cascade remove os valores
        db.session.commit()
        return jsonify({"ok": True}), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em DELETE /rh/convencoes/<id>")
        return jsonify({"erro": "Erro ao remover convenção", "detalhe": str(e)}), 500


# ---------------------------------------------------------------- funcionários

@rh_bp.route('/funcionarios', methods=['GET'])
@jwt_required()
def listar_funcionarios():
    try:
        current_user = get_current_user()
        q = Funcionario.query.options(
            joinedload(Funcionario.categoria), joinedload(Funcionario.obra)
        )
        obra_id_raw = request.args.get('obra_id')
        status = request.args.get('status')
        if obra_id_raw is not None and obra_id_raw != '':
            if obra_id_raw in ('null', 'sem', 'sem_obra'):
                if not user_has_access_to_obra(current_user, None):
                    return jsonify({"erro": "Acesso negado"}), 403
                q = q.filter(Funcionario.obra_id.is_(None))
            else:
                obra_id, err = _parse_int_arg('obra_id', obra_id_raw)
                if err:
                    return err
                if not user_has_access_to_obra(current_user, obra_id):
                    return jsonify({"erro": "Acesso negado a esta obra"}), 403
                q = q.filter(Funcionario.obra_id == obra_id)
        else:
            q = _restringir_por_obra(q, Funcionario.obra_id, current_user)
        if status:
            q = q.filter(Funcionario.status == status)
        funcs = q.order_by(Funcionario.nome).all()
        piso_lookup = rh_service.piso_vigente_batch(
            (f.categoria_id, f.obra.uf if f.obra else None) for f in funcs
        )
        return jsonify([f.to_dict(piso_lookup=piso_lookup) for f in funcs]), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/funcionarios")
        return jsonify({"erro": "Erro ao listar funcionários", "detalhe": str(e)}), 500


@rh_bp.route('/funcionarios/piso-sugerido', methods=['GET'])
@jwt_required()
def piso_sugerido():
    """Piso pra pré-preencher o salário no form (via UF da obra)."""
    try:
        categoria_id = request.args.get('categoria_id', type=int)
        obra_id = request.args.get('obra_id', type=int)
        uf = None
        if obra_id:
            from models.obra import Obra
            obra = db.session.get(Obra, obra_id)
            uf = obra.uf if obra else None
        piso = rh_service.piso_vigente(categoria_id, uf) if categoria_id else None
        return jsonify({"piso_sugerido": piso, "uf": uf}), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/funcionarios/piso-sugerido")
        return jsonify({"erro": "Erro ao sugerir piso", "detalhe": str(e)}), 500


@rh_bp.route('/funcionarios', methods=['POST'])
@jwt_required()
def criar_funcionario():
    try:
        dados = request.get_json() or {}
        nome = (dados.get('nome') or '').strip()
        categoria_id = dados.get('categoria_id')
        if not nome or not categoria_id:
            return jsonify({"erro": "nome e categoria_id são obrigatórios"}), 400
        categoria_id = int(categoria_id)
        if not db.session.get(CategoriaMO, categoria_id):
            return jsonify({"erro": "Categoria não encontrada"}), 400

        obra_id = int(dados['obra_id']) if dados.get('obra_id') else None
        obra = None
        if obra_id is not None:
            obra = db.session.get(Obra, obra_id)
            if not obra:
                return jsonify({"erro": "Obra não encontrada"}), 400

        current_user = get_current_user()
        if not user_has_access_to_obra(current_user, obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403

        salario = _to_num(dados.get('salario'))
        if salario is None:
            # fallback: piso vigente da categoria pela UF da obra
            uf = obra.uf if obra else None
            salario = rh_service.piso_vigente(categoria_id, uf) or 0

        func = Funcionario(
            nome=nome,
            cpf=(dados.get('cpf') or '').strip() or None,
            categoria_id=categoria_id,
            obra_id=obra_id,
            salario=salario,
            data_admissao=_parse_date(dados.get('data_admissao')),
            status=dados.get('status') or 'ativo',
        )
        db.session.add(func)
        db.session.commit()
        return jsonify(func.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em POST /rh/funcionarios")
        return jsonify({"erro": "Erro ao criar funcionário", "detalhe": str(e)}), 500


@rh_bp.route('/funcionarios/importar', methods=['POST'])
@jwt_required()
def importar_funcionarios():
    """Importa funcionários de planilha (.xlsx/.csv).

    Colunas (cabeçalho, case-insensitive): nome (obrig.), categoria (obrig.,
    nome — cria se não existir), cpf, obra (id ou nome), salario, admissao.
    Linhas inválidas são ignoradas e reportadas. Erros de validação = 400."""
    try:
        arquivo = request.files.get('arquivo') or request.files.get('file')
        if not arquivo:
            return jsonify({"erro": "arquivo (.xlsx/.csv) é obrigatório"}), 400

        nome_arq = (arquivo.filename or '').lower()
        linhas = []
        if nome_arq.endswith('.csv'):
            texto = arquivo.read().decode('utf-8-sig', errors='replace')
            linhas = list(csv.DictReader(io.StringIO(texto)))
        elif nome_arq.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(arquivo.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                cab = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
                for r in rows[1:]:
                    linhas.append({cab[i]: r[i] for i in range(min(len(cab), len(r)))})
        else:
            return jsonify({"erro": "formato não suportado (use .xlsx ou .csv)"}), 400

        current_user = get_current_user()
        obras_por_nome = {o.nome.strip().lower(): o for o in Obra.query.all()}
        criados, erros = 0, []

        def _erro(idx, motivo):
            if len(erros) < 30:
                erros.append({'linha': idx + 2, 'motivo': motivo})  # +2 = 1-based + cabeçalho

        for idx, ln in enumerate(linhas):
            ln = {str(k).strip().lower(): v for k, v in ln.items() if k}
            nome = str(ln.get('nome') or '').strip()
            cat_nome = str(ln.get('categoria') or '').strip()
            if not nome:
                _erro(idx, 'nome vazio')
                continue
            if not cat_nome:
                _erro(idx, 'categoria vazia')
                continue

            obra = None
            obra_ref = ln.get('obra') or ln.get('obra_id')
            if obra_ref not in (None, ''):
                try:
                    obra = db.session.get(Obra, int(obra_ref))
                except (TypeError, ValueError):
                    obra = obras_por_nome.get(str(obra_ref).strip().lower())
                if not obra:
                    _erro(idx, f'obra "{obra_ref}" não encontrada')
                    continue
            obra_id = obra.id if obra else None
            if not user_has_access_to_obra(current_user, obra_id):
                _erro(idx, 'sem acesso à obra')
                continue

            categoria = _resolver_categoria(cat_nome)
            salario = _to_num(ln.get('salario'))
            if salario is None:
                uf = obra.uf if obra else None
                salario = rh_service.piso_vigente(categoria.id, uf) or 0

            db.session.add(Funcionario(
                nome=nome,
                cpf=str(ln.get('cpf') or '').strip() or None,
                categoria_id=categoria.id,
                obra_id=obra_id,
                salario=salario,
                data_admissao=_parse_date(ln.get('admissao') or ln.get('data_admissao')),
                status='ativo',
            ))
            criados += 1

        db.session.commit()
        return jsonify({"criados": criados, "ignorados": len(linhas) - criados,
                        "erros": erros}), 201
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em POST /rh/funcionarios/importar")
        return jsonify({"erro": "Erro ao importar funcionários", "detalhe": str(e)}), 500


@rh_bp.route('/funcionarios/<int:func_id>', methods=['GET'])
@jwt_required()
def obter_funcionario(func_id):
    try:
        func = db.session.get(Funcionario, func_id)
        if not func:
            return jsonify({"erro": "Funcionário não encontrado"}), 404
        if not user_has_access_to_obra(get_current_user(), func.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        return jsonify(func.to_dict()), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/funcionarios/<id>")
        return jsonify({"erro": "Erro ao obter funcionário", "detalhe": str(e)}), 500


@rh_bp.route('/funcionarios/<int:func_id>', methods=['PUT'])
@jwt_required()
def editar_funcionario(func_id):
    try:
        func = db.session.get(Funcionario, func_id)
        if not func:
            return jsonify({"erro": "Funcionário não encontrado"}), 404
        current_user = get_current_user()
        if not user_has_access_to_obra(current_user, func.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        dados = request.get_json() or {}
        if 'nome' in dados and dados['nome']:
            func.nome = dados['nome'].strip()
        if 'cpf' in dados:
            func.cpf = (dados.get('cpf') or '').strip() or None
        if dados.get('categoria_id'):
            novo_cat_id = int(dados['categoria_id'])
            if not db.session.get(CategoriaMO, novo_cat_id):
                return jsonify({"erro": "Categoria não encontrada"}), 400
            func.categoria_id = novo_cat_id
        if 'obra_id' in dados:
            novo_obra_id = int(dados['obra_id']) if dados.get('obra_id') else None
            if novo_obra_id is not None and not db.session.get(Obra, novo_obra_id):
                return jsonify({"erro": "Obra não encontrada"}), 400
            if not user_has_access_to_obra(current_user, novo_obra_id):
                return jsonify({"erro": "Acesso negado a esta obra"}), 403
            func.obra_id = novo_obra_id
        if dados.get('salario') is not None:
            novo = _to_num(dados.get('salario'))
            if novo is not None:
                func.salario = novo
        if 'data_admissao' in dados:
            func.data_admissao = _parse_date(dados.get('data_admissao'))
        if 'data_demissao' in dados:
            func.data_demissao = _parse_date(dados.get('data_demissao'))
        if dados.get('status') in ('ativo', 'inativo', 'demitido'):
            func.status = dados['status']
        db.session.commit()
        return jsonify(func.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em PUT /rh/funcionarios/<id>")
        return jsonify({"erro": "Erro ao editar funcionário", "detalhe": str(e)}), 500


@rh_bp.route('/funcionarios/<int:func_id>/obra', methods=['PATCH'])
@jwt_required()
def migrar_obra_funcionario(func_id):
    """Migra o funcionário de obra (body {obra_id}). Pagamentos passados mantêm
    o snapshot da obra antiga."""
    try:
        func = db.session.get(Funcionario, func_id)
        if not func:
            return jsonify({"erro": "Funcionário não encontrado"}), 404
        current_user = get_current_user()
        if not user_has_access_to_obra(current_user, func.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        dados = request.get_json() or {}
        obra_id = dados.get('obra_id')
        novo_obra_id = int(obra_id) if obra_id else None
        if novo_obra_id is not None and not db.session.get(Obra, novo_obra_id):
            return jsonify({"erro": "Obra não encontrada"}), 400
        if not user_has_access_to_obra(current_user, novo_obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        func.obra_id = novo_obra_id
        db.session.commit()
        return jsonify(func.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em PATCH /rh/funcionarios/<id>/obra")
        return jsonify({"erro": "Erro ao migrar obra", "detalhe": str(e)}), 500


@rh_bp.route('/funcionarios/<int:func_id>', methods=['DELETE'])
@jwt_required()
def inativar_funcionario(func_id):
    """Inativa (status), não apaga."""
    try:
        func = db.session.get(Funcionario, func_id)
        if not func:
            return jsonify({"erro": "Funcionário não encontrado"}), 404
        if not user_has_access_to_obra(get_current_user(), func.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        func.status = 'inativo'
        db.session.commit()
        return jsonify(func.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em DELETE /rh/funcionarios/<id>")
        return jsonify({"erro": "Erro ao inativar funcionário", "detalhe": str(e)}), 500


# ---------------------------------------------------------------- pagamentos

@rh_bp.route('/pagamentos', methods=['GET'])
@jwt_required()
def listar_pagamentos():
    try:
        current_user = get_current_user()
        q = PagamentoSalario.query.options(
            joinedload(PagamentoSalario.funcionario), joinedload(PagamentoSalario.obra)
        )
        if request.args.get('competencia'):
            q = q.filter(PagamentoSalario.competencia == request.args['competencia'])
        obra_id_raw = request.args.get('obra_id')
        if obra_id_raw:
            obra_id, err = _parse_int_arg('obra_id', obra_id_raw)
            if err:
                return err
            if not user_has_access_to_obra(current_user, obra_id):
                return jsonify({"erro": "Acesso negado a esta obra"}), 403
            q = q.filter(PagamentoSalario.obra_id == obra_id)
        else:
            q = _restringir_por_obra(q, PagamentoSalario.obra_id, current_user)
        func_id_raw = request.args.get('funcionario_id')
        if func_id_raw:
            func_id, err = _parse_int_arg('funcionario_id', func_id_raw)
            if err:
                return err
            q = q.filter(PagamentoSalario.funcionario_id == func_id)
        pags = q.order_by(PagamentoSalario.data_pagamento.desc()).all()
        return jsonify([p.to_dict() for p in pags]), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/pagamentos")
        return jsonify({"erro": "Erro ao listar pagamentos", "detalhe": str(e)}), 500


@rh_bp.route('/pagamentos', methods=['POST'])
@jwt_required()
def criar_pagamento():
    """Cria pagamento. obra_id é SNAPSHOT: copiado do funcionário no POST."""
    try:
        if request.content_type and 'multipart' in request.content_type:
            dados = request.form
            arquivo = request.files.get('arquivo') or request.files.get('comprovante')
        else:
            dados = request.get_json() or {}
            arquivo = None

        funcionario_id = dados.get('funcionario_id')
        competencia = (dados.get('competencia') or '').strip()
        tipo = (dados.get('tipo') or 'salario').strip()
        valor = _to_num(dados.get('valor'))
        data_pag = _parse_date(dados.get('data_pagamento'))

        if not funcionario_id or not competencia or valor is None or not data_pag:
            return jsonify({"erro": "funcionario_id, competencia, valor e data_pagamento são obrigatórios"}), 400
        if tipo not in _PAG_TIPOS:
            return jsonify({"erro": f"tipo inválido (use {sorted(_PAG_TIPOS)})"}), 400

        func = db.session.get(Funcionario, int(funcionario_id))
        if not func:
            return jsonify({"erro": "Funcionário não encontrado"}), 404
        if not user_has_access_to_obra(get_current_user(), func.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403

        comprovante_url = None
        anexo_falhou = False
        if arquivo:
            try:
                comprovante_url = storage_service.upload_arquivo(arquivo, 'comprovantes')
            except Exception as e:
                anexo_falhou = True
                logger.exception("Upload do comprovante falhou (segue sem): %s", e)

        pag = PagamentoSalario(
            funcionario_id=func.id,
            competencia=competencia,
            tipo=tipo,
            valor=valor,
            data_pagamento=data_pag,
            obra_id=func.obra_id,           # SNAPSHOT
            comprovante_url=comprovante_url,
            observacao=(dados.get('observacao') or '').strip() or None,
        )
        db.session.add(pag)
        db.session.commit()
        resp = pag.to_dict()
        if anexo_falhou:
            resp['aviso'] = "Pagamento salvo, mas o comprovante não pôde ser enviado."
        return jsonify(resp), 201
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em POST /rh/pagamentos")
        return jsonify({"erro": "Erro ao criar pagamento", "detalhe": str(e)}), 500


@rh_bp.route('/pagamentos/<int:pag_id>', methods=['GET'])
@jwt_required()
def obter_pagamento(pag_id):
    try:
        pag = db.session.get(PagamentoSalario, pag_id)
        if not pag:
            return jsonify({"erro": "Pagamento não encontrado"}), 404
        if not user_has_access_to_obra(get_current_user(), pag.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        return jsonify(pag.to_dict()), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/pagamentos/<id>")
        return jsonify({"erro": "Erro ao obter pagamento", "detalhe": str(e)}), 500


@rh_bp.route('/pagamentos/<int:pag_id>', methods=['PUT'])
@jwt_required()
def editar_pagamento(pag_id):
    try:
        pag = db.session.get(PagamentoSalario, pag_id)
        if not pag:
            return jsonify({"erro": "Pagamento não encontrado"}), 404
        if not user_has_access_to_obra(get_current_user(), pag.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        dados = request.get_json() or {}
        if dados.get('competencia'):
            pag.competencia = dados['competencia'].strip()
        if dados.get('tipo') in _PAG_TIPOS:
            pag.tipo = dados['tipo']
        if dados.get('valor') is not None:
            novo = _to_num(dados.get('valor'))
            if novo is not None:
                pag.valor = novo
        if dados.get('data_pagamento'):
            pag.data_pagamento = _parse_date(dados['data_pagamento']) or pag.data_pagamento
        if 'observacao' in dados:
            pag.observacao = (dados.get('observacao') or '').strip() or None
        db.session.commit()
        return jsonify(pag.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em PUT /rh/pagamentos/<id>")
        return jsonify({"erro": "Erro ao editar pagamento", "detalhe": str(e)}), 500


@rh_bp.route('/pagamentos/<int:pag_id>', methods=['DELETE'])
@jwt_required()
def remover_pagamento(pag_id):
    try:
        pag = db.session.get(PagamentoSalario, pag_id)
        if not pag:
            return jsonify({"erro": "Pagamento não encontrado"}), 404
        if not user_has_access_to_obra(get_current_user(), pag.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        db.session.delete(pag)
        db.session.commit()
        return jsonify({"ok": True}), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em DELETE /rh/pagamentos/<id>")
        return jsonify({"erro": "Erro ao remover pagamento", "detalhe": str(e)}), 500


# ---------------------------------------------------------------- encargos

@rh_bp.route('/encargos', methods=['GET'])
@jwt_required()
def listar_encargos():
    try:
        current_user = get_current_user()
        q = Encargo.query.options(
            joinedload(Encargo.obra), joinedload(Encargo.funcionario)
        )
        if request.args.get('competencia'):
            q = q.filter(Encargo.competencia == request.args['competencia'])
        if request.args.get('tipo'):
            q = q.filter(Encargo.tipo == request.args['tipo'])
        obra_id_raw = request.args.get('obra_id')
        if obra_id_raw:
            obra_id, err = _parse_int_arg('obra_id', obra_id_raw)
            if err:
                return err
            if not user_has_access_to_obra(current_user, obra_id):
                return jsonify({"erro": "Acesso negado a esta obra"}), 403
            q = q.filter(Encargo.obra_id == obra_id)
        else:
            q = _restringir_por_obra(q, Encargo.obra_id, current_user)
        encargos = q.order_by(Encargo.vencimento.desc().nullslast()).all()
        return jsonify([e.to_dict() for e in encargos]), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/encargos")
        return jsonify({"erro": "Erro ao listar encargos", "detalhe": str(e)}), 500


@rh_bp.route('/encargos', methods=['POST'])
@jwt_required()
def criar_encargo():
    try:
        if request.content_type and 'multipart' in request.content_type:
            dados = request.form
            arquivo = request.files.get('arquivo') or request.files.get('file')
        else:
            dados = request.get_json() or {}
            arquivo = None

        tipo = (dados.get('tipo') or '').strip()
        competencia = (dados.get('competencia') or '').strip()
        valor = _to_num(dados.get('valor'))
        if tipo not in _ENCARGO_TIPOS or not competencia or valor is None:
            return jsonify({"erro": f"tipo (∈ {sorted(_ENCARGO_TIPOS)}), competencia e valor são obrigatórios"}), 400

        obra_id = int(dados['obra_id']) if dados.get('obra_id') else None
        if obra_id is not None and not db.session.get(Obra, obra_id):
            return jsonify({"erro": "Obra não encontrada"}), 400
        funcionario_id = int(dados['funcionario_id']) if dados.get('funcionario_id') else None
        if funcionario_id is not None and not db.session.get(Funcionario, funcionario_id):
            return jsonify({"erro": "Funcionário não encontrado"}), 400
        if not user_has_access_to_obra(get_current_user(), obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403

        arquivo_url = None
        anexo_falhou = False
        if arquivo:
            try:
                arquivo_url = storage_service.upload_arquivo(arquivo, 'guias')
            except Exception as e:
                anexo_falhou = True
                logger.exception("Upload da guia falhou (segue sem): %s", e)

        enc = Encargo(
            tipo=tipo,
            competencia=competencia,
            vencimento=_parse_date(dados.get('vencimento')),
            data_pagamento=_parse_date(dados.get('data_pagamento')),
            valor=valor,
            arquivo_url=arquivo_url,
            obra_id=obra_id,
            funcionario_id=funcionario_id,
            observacao=(dados.get('observacao') or '').strip() or None,
        )
        db.session.add(enc)
        db.session.commit()
        resp = enc.to_dict()
        if anexo_falhou:
            resp['aviso'] = "Encargo salvo, mas a guia (PDF) não pôde ser enviada."
        return jsonify(resp), 201
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em POST /rh/encargos")
        return jsonify({"erro": "Erro ao criar encargo", "detalhe": str(e)}), 500


@rh_bp.route('/encargos/sugestao', methods=['GET'])
@jwt_required()
def sugestao_encargo():
    """Valor sugerido do encargo a partir da folha da competência.

    ?tipo=fgts|inss_darf|esocial_dae &competencia=YYYY-MM &obra_id=<id|sem>
    Base = soma dos pagamentos de salário (tipo 'salario') da competência,
    no escopo da obra (ou geral). Percentuais FEDERAIS (não variam por
    região — o que varia por região é o piso da CCT, já coberto):
    FGTS 8%; INSS patronal 20% (sem RAT/terceiros — confira com a
    contabilidade); eSocial/DAE não tem percentual único → só a base."""
    _PERCENTUAIS = {'fgts': 0.08, 'inss_darf': 0.20, 'esocial_dae': None}
    try:
        tipo = (request.args.get('tipo') or '').strip()
        if tipo not in _ENCARGO_TIPOS:
            return jsonify({"erro": f"tipo inválido (use {sorted(_ENCARGO_TIPOS)})"}), 400
        competencia = (request.args.get('competencia') or '').strip()
        if not competencia or len(competencia) != 7:
            return jsonify({"erro": "competencia é obrigatória (YYYY-MM)"}), 400

        user = get_current_user()
        query = (PagamentoSalario.query
                 .filter(PagamentoSalario.competencia == competencia,
                         PagamentoSalario.tipo == 'salario'))
        obra_arg = request.args.get('obra_id')
        if obra_arg in ('sem', 'null', 'sem_obra'):
            query = query.filter(PagamentoSalario.obra_id.is_(None))
        elif obra_arg:
            obra_id, erro = _parse_int_arg('obra_id', obra_arg)
            if erro:
                return erro
            if not user_has_access_to_obra(user, obra_id):
                return jsonify({"erro": "Acesso negado a esta obra"}), 403
            query = query.filter(PagamentoSalario.obra_id == obra_id)
        else:
            query = _restringir_por_obra(query, PagamentoSalario.obra_id, user)

        base = float(query.with_entities(
            db.func.coalesce(db.func.sum(PagamentoSalario.valor), 0)).scalar() or 0)
        pct = _PERCENTUAIS.get(tipo)
        return jsonify({
            'tipo': tipo,
            'competencia': competencia,
            'base_folha': round(base, 2),
            'percentual': pct,
            'valor_sugerido': round(base * pct, 2) if pct is not None else None,
            'nota': ('FGTS 8% sobre a folha da competência' if tipo == 'fgts'
                     else 'INSS patronal 20% (sem RAT/terceiros) — confirme com a contabilidade'
                     if tipo == 'inss_darf'
                     else 'Sem percentual único — informe o valor da guia'),
        }), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/encargos/sugestao")
        return jsonify({"erro": "Erro ao calcular sugestão", "detalhe": str(e)}), 500


@rh_bp.route('/encargos/importar', methods=['POST'])
@jwt_required()
def importar_encargos():
    """Importa planilha da contabilidade (.xlsx/.csv). Colunas:
    tipo, competencia, vencimento, valor, obra? (nome ou id opcional)."""
    try:
        arquivo = request.files.get('arquivo') or request.files.get('file')
        if not arquivo:
            return jsonify({"erro": "arquivo (.xlsx/.csv) é obrigatório"}), 400

        nome = (arquivo.filename or '').lower()
        linhas = []
        if nome.endswith('.csv'):
            texto = arquivo.read().decode('utf-8-sig', errors='replace')
            linhas = list(csv.DictReader(io.StringIO(texto)))
        elif nome.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(arquivo.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                cab = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
                for r in rows[1:]:
                    linhas.append({cab[i]: r[i] for i in range(min(len(cab), len(r)))})
        else:
            return jsonify({"erro": "formato não suportado (use .xlsx ou .csv)"}), 400

        current_user = get_current_user()
        criados, ignorados = 0, 0
        for ln in linhas:
            tipo = str(ln.get('tipo') or '').strip().lower()
            competencia = str(ln.get('competencia') or '').strip()
            valor = _to_num(ln.get('valor'))
            if tipo not in _ENCARGO_TIPOS or not competencia or valor is None:
                ignorados += 1
                continue
            obra_id = ln.get('obra') or ln.get('obra_id')
            try:
                obra_id = int(obra_id) if obra_id not in (None, '') else None
            except Exception:
                obra_id = None
            if not user_has_access_to_obra(current_user, obra_id):
                ignorados += 1
                continue
            db.session.add(Encargo(
                tipo=tipo,
                competencia=competencia,
                vencimento=_parse_date(ln.get('vencimento')),
                valor=valor,
                obra_id=obra_id,
            ))
            criados += 1

        db.session.commit()
        return jsonify({"criados": criados, "ignorados": ignorados}), 201
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em POST /rh/encargos/importar")
        return jsonify({"erro": "Erro ao importar encargos", "detalhe": str(e)}), 500


@rh_bp.route('/encargos/<int:enc_id>', methods=['GET'])
@jwt_required()
def obter_encargo(enc_id):
    try:
        enc = db.session.get(Encargo, enc_id)
        if not enc:
            return jsonify({"erro": "Encargo não encontrado"}), 404
        if not user_has_access_to_obra(get_current_user(), enc.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        return jsonify(enc.to_dict()), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/encargos/<id>")
        return jsonify({"erro": "Erro ao obter encargo", "detalhe": str(e)}), 500


@rh_bp.route('/encargos/<int:enc_id>', methods=['PUT'])
@jwt_required()
def editar_encargo(enc_id):
    try:
        enc = db.session.get(Encargo, enc_id)
        if not enc:
            return jsonify({"erro": "Encargo não encontrado"}), 404
        current_user = get_current_user()
        if not user_has_access_to_obra(current_user, enc.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        dados = request.get_json() or {}
        if dados.get('tipo') in _ENCARGO_TIPOS:
            enc.tipo = dados['tipo']
        if dados.get('competencia'):
            enc.competencia = dados['competencia'].strip()
        if 'vencimento' in dados:
            enc.vencimento = _parse_date(dados.get('vencimento'))
        if 'data_pagamento' in dados:
            enc.data_pagamento = _parse_date(dados.get('data_pagamento'))
        if dados.get('valor') is not None:
            novo = _to_num(dados.get('valor'))
            if novo is not None:
                enc.valor = novo
        if 'obra_id' in dados:
            novo_obra_id = int(dados['obra_id']) if dados.get('obra_id') else None
            if novo_obra_id is not None and not db.session.get(Obra, novo_obra_id):
                return jsonify({"erro": "Obra não encontrada"}), 400
            if not user_has_access_to_obra(current_user, novo_obra_id):
                return jsonify({"erro": "Acesso negado a esta obra"}), 403
            enc.obra_id = novo_obra_id
        if 'observacao' in dados:
            enc.observacao = (dados.get('observacao') or '').strip() or None
        db.session.commit()
        return jsonify(enc.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em PUT /rh/encargos/<id>")
        return jsonify({"erro": "Erro ao editar encargo", "detalhe": str(e)}), 500


@rh_bp.route('/encargos/<int:enc_id>', methods=['DELETE'])
@jwt_required()
def remover_encargo(enc_id):
    try:
        enc = db.session.get(Encargo, enc_id)
        if not enc:
            return jsonify({"erro": "Encargo não encontrado"}), 404
        if not user_has_access_to_obra(get_current_user(), enc.obra_id):
            return jsonify({"erro": "Acesso negado a esta obra"}), 403
        db.session.delete(enc)
        db.session.commit()
        return jsonify({"ok": True}), 200
    except Exception as e:
        db.session.rollback()
        logger.exception("Erro em DELETE /rh/encargos/<id>")
        return jsonify({"erro": "Erro ao remover encargo", "detalhe": str(e)}), 500


# ---------------------------------------------------------------- arquivo / dashboard

@rh_bp.route('/arquivo/<tipo>/<int:item_id>', methods=['GET'])
@jwt_required()
def obter_arquivo(tipo, item_id):
    """Retorna signed URL do arquivo (convencao | comprovante | guia) sob auth.

    Convenção é dado global (por UF), sem obra_id — não é obra-scoped.
    Comprovante/guia herdam o obra_id do registro (pagamento/encargo)."""
    try:
        path = None
        if tipo == 'convencao':
            obj = db.session.get(ConvencaoColetiva, item_id)
            path = obj.arquivo_url if obj else None
        elif tipo == 'comprovante':
            obj = db.session.get(PagamentoSalario, item_id)
            if obj and not user_has_access_to_obra(get_current_user(), obj.obra_id):
                return jsonify({"erro": "Acesso negado a esta obra"}), 403
            path = obj.comprovante_url if obj else None
        elif tipo == 'guia':
            obj = db.session.get(Encargo, item_id)
            if obj and not user_has_access_to_obra(get_current_user(), obj.obra_id):
                return jsonify({"erro": "Acesso negado a esta obra"}), 403
            path = obj.arquivo_url if obj else None
        else:
            return jsonify({"erro": "tipo inválido (use convencao|comprovante|guia)"}), 400

        if not path:
            return jsonify({"erro": "Arquivo não encontrado"}), 404
        return jsonify({"url": storage_service.signed_url(path)}), 200
    except RuntimeError as e:
        return jsonify({"erro": str(e)}), 503
    except Exception as e:
        logger.exception("Erro em GET /rh/arquivo/<tipo>/<id>")
        return jsonify({"erro": "Erro ao obter arquivo", "detalhe": str(e)}), 500


@rh_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def obter_dashboard():
    try:
        competencia = request.args.get('competencia')
        if not competencia:
            competencia = date.today().strftime('%Y-%m')
        return jsonify(rh_service.dashboard(competencia)), 200
    except Exception as e:
        logger.exception("Erro em GET /rh/dashboard")
        return jsonify({"erro": "Erro ao montar dashboard", "detalhe": str(e)}), 500
