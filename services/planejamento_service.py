import csv
import io
import math
import re
import unicodedata
import zipfile
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

from sqlalchemy import and_, or_

from extensions import db
from models.cronograma_obra import CronogramaObra
from models.orcamento_eng_etapa import OrcamentoEngEtapa
from models.orcamento_eng_item import OrcamentoEngItem
from models.planejamento_atividade import (
    ORIGENS_PLANEJAMENTO,
    PRIORIDADES_PLANEJAMENTO,
    STATUS_PLANEJAMENTO,
    PlanejamentoAtividade,
)


MAX_PLANILHA_BYTES = 2 * 1024 * 1024
MAX_PLANILHA_UNCOMPRESSED_BYTES = 12 * 1024 * 1024
MAX_PLANILHA_ROWS = 500
MAX_PLANILHA_COLUMNS = 30


class PlanejamentoValidationError(ValueError):
    def __init__(self, message, field=None, details=None):
        super().__init__(message)
        self.field = field
        self.details = details or []


def parse_iso_date(value, field, required=False):
    if value in (None, ''):
        if required:
            raise PlanejamentoValidationError(f'{field} é obrigatório.', field)
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise PlanejamentoValidationError(
            f'{field} deve usar o formato AAAA-MM-DD.', field
        ) from exc


def parse_decimal(value, field, minimum=0, maximum=Decimal('1000000000000')):
    if value in (None, ''):
        return Decimal('0')
    try:
        number = Decimal(str(value).strip().replace(',', '.'))
    except (InvalidOperation, AttributeError, ValueError) as exc:
        raise PlanejamentoValidationError(f'{field} deve ser numérico.', field) from exc
    if not number.is_finite() or number < Decimal(str(minimum)) or number > maximum:
        raise PlanejamentoValidationError(
            f'{field} deve estar entre {minimum} e {maximum}.', field
        )
    return number.quantize(Decimal('0.001'))


def clean_text(value, field, maximum, required=False):
    if value is None:
        value = ''
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', str(value)).strip()
    if required and not text:
        raise PlanejamentoValidationError(f'{field} é obrigatório.', field)
    if len(text) > maximum:
        raise PlanejamentoValidationError(
            f'{field} deve ter no máximo {maximum} caracteres.', field
        )
    return text or None


def normalize_enum(value, field, allowed, default=None):
    if value in (None, ''):
        return default
    normalized = str(value).strip().lower()
    if normalized not in allowed:
        raise PlanejamentoValidationError(
            f'{field} inválido. Valores aceitos: {", ".join(allowed)}.', field
        )
    return normalized


def normalize_activity_payload(data, origem='manual', partial=False):
    if not isinstance(data, dict):
        raise PlanejamentoValidationError('O corpo da requisição deve ser um objeto JSON.')
    if origem not in ORIGENS_PLANEJAMENTO:
        raise PlanejamentoValidationError('Origem de atividade inválida.', 'origem')

    result = {}
    text_fields = {
        'titulo': (240, True),
        'descricao': (4000, False),
        'etapa_nome': (200, False),
        'responsavel': (160, False),
        'equipe': (160, False),
        'unidade': (24, False),
        'observacoes': (4000, False),
    }
    for field, (limit, required) in text_fields.items():
        if field in data or (field == 'titulo' and not partial):
            result[field] = clean_text(
                data.get(field), field, limit, required=required and not partial
            )

    if not partial:
        result.setdefault('unidade', 'un')
        result['origem'] = origem

    if 'status' in data or not partial:
        result['status'] = normalize_enum(
            data.get('status'), 'status', STATUS_PLANEJAMENTO, 'a_planejar'
        )
    if 'prioridade' in data or not partial:
        result['prioridade'] = normalize_enum(
            data.get('prioridade'),
            'prioridade',
            PRIORIDADES_PLANEJAMENTO,
            'normal',
        )

    if 'data_inicio' in data or not partial:
        result['data_inicio'] = parse_iso_date(data.get('data_inicio'), 'data_inicio')
    if 'data_fim' in data or not partial:
        result['data_fim'] = parse_iso_date(data.get('data_fim'), 'data_fim')
    if (
        result.get('data_inicio')
        and result.get('data_fim')
        and result['data_fim'] < result['data_inicio']
    ):
        raise PlanejamentoValidationError(
            'data_fim não pode ser anterior a data_inicio.', 'data_fim'
        )

    if 'quantidade_planejada' in data or not partial:
        result['quantidade_planejada'] = parse_decimal(
            data.get('quantidade_planejada'), 'quantidade_planejada'
        )

    if 'cronograma_id' in data:
        try:
            cronograma_id = int(data.get('cronograma_id')) if data.get('cronograma_id') else None
        except (TypeError, ValueError) as exc:
            raise PlanejamentoValidationError(
                'cronograma_id deve ser um inteiro.', 'cronograma_id'
            ) from exc
        result['cronograma_id'] = cronograma_id

    return result


def validate_cronograma_belongs_to_obra(cronograma_id, obra_id):
    if cronograma_id is None:
        return None
    cronograma = db.session.get(CronogramaObra, cronograma_id)
    if not cronograma or cronograma.obra_id != obra_id:
        raise PlanejamentoValidationError(
            'Cronograma inexistente ou não pertence à obra.', 'cronograma_id'
        )
    return cronograma


def apply_activity_fields(atividade, fields):
    for field, value in fields.items():
        setattr(atividade, field, value)
    atividade.versao = (atividade.versao or 0) + 1 if atividade.id else 1
    return atividade


def automatic_status(atividade):
    if any(restricao.status == 'aberta' for restricao in atividade.restricoes):
        return 'impedido'
    planned = Decimal(str(atividade.quantidade_planejada or 0))
    executed = Decimal(str(atividade.quantidade_executada or 0))
    if planned > 0 and executed >= planned:
        return 'concluido'
    if executed > 0:
        return 'em_andamento'
    if atividade.data_inicio and atividade.data_fim:
        return 'pronto'
    return 'a_planejar'


def activity_overlaps_period(start_date, end_date):
    return and_(
        or_(PlanejamentoAtividade.data_inicio.is_(None), PlanejamentoAtividade.data_inicio <= end_date),
        or_(PlanejamentoAtividade.data_fim.is_(None), PlanejamentoAtividade.data_fim >= start_date),
    )


def start_of_week(value=None):
    current = parse_iso_date(value, 'semana_inicio') if value else date.today()
    return current - timedelta(days=current.weekday())


def summarize_activities(activities):
    counts = {status: 0 for status in STATUS_PLANEJAMENTO}
    planned_total = Decimal('0')
    executed_total = Decimal('0')
    open_constraints = 0
    for activity in activities:
        counts[activity.status] = counts.get(activity.status, 0) + 1
        planned_total += Decimal(str(activity.quantidade_planejada or 0))
        executed_total += Decimal(str(activity.quantidade_executada or 0))
        open_constraints += sum(1 for item in activity.restricoes if item.status == 'aberta')
    completed = counts.get('concluido', 0)
    committed = len([a for a in activities if a.status != 'a_planejar'])
    return {
        'total': len(activities),
        'por_status': counts,
        'quantidade_planejada': float(planned_total),
        'quantidade_executada': float(executed_total),
        'restricoes_abertas': open_constraints,
        'confiabilidade': round(completed / committed * 100, 1) if committed else 0,
    }


def import_budget_items(obra_id, item_ids, defaults, complements, user_id):
    if not isinstance(item_ids, list) or not item_ids or len(item_ids) > 200:
        raise PlanejamentoValidationError(
            'item_ids deve conter entre 1 e 200 itens.', 'item_ids'
        )
    try:
        normalized_ids = list(dict.fromkeys(int(item_id) for item_id in item_ids))
    except (TypeError, ValueError) as exc:
        raise PlanejamentoValidationError(
            'Todos os item_ids devem ser inteiros.', 'item_ids'
        ) from exc

    items = (
        OrcamentoEngItem.query
        .join(OrcamentoEngEtapa, OrcamentoEngItem.etapa_id == OrcamentoEngEtapa.id)
        .filter(
            OrcamentoEngEtapa.obra_id == obra_id,
            OrcamentoEngItem.id.in_(normalized_ids),
        )
        .all()
    )
    found_ids = {item.id for item in items}
    missing = sorted(set(normalized_ids) - found_ids)
    if missing:
        raise PlanejamentoValidationError(
            'Há itens inexistentes ou pertencentes a outra obra.',
            'item_ids',
            details=missing,
        )

    existing = {
        row.orcamento_item_id
        for row in PlanejamentoAtividade.query.filter(
            PlanejamentoAtividade.obra_id == obra_id,
            PlanejamentoAtividade.orcamento_item_id.in_(normalized_ids),
        ).all()
    }
    base_defaults = normalize_activity_payload(defaults or {}, origem='orcamento', partial=True)
    created = []
    skipped = []
    for item in items:
        if item.id in existing:
            skipped.append(item.id)
            continue
        item_complement = {}
        if isinstance(complements, dict):
            item_complement = complements.get(str(item.id), complements.get(item.id, {})) or {}
        complement_fields = normalize_activity_payload(
            item_complement, origem='orcamento', partial=True
        )
        fields = {
            **base_defaults,
            **complement_fields,
            'titulo': item.descricao,
            'descricao': f'Importado do orçamento: {item.codigo or "sem código"}',
            'etapa_nome': item.etapa.nome if item.etapa else None,
            'quantidade_planejada': parse_decimal(
                item.quantidade or 0, 'quantidade_planejada'
            ),
            'unidade': clean_text(item.unidade or 'un', 'unidade', 24) or 'un',
            'origem': 'orcamento',
        }
        if not fields.get('data_inicio') or not fields.get('data_fim'):
            fields['status'] = 'a_planejar'
        validate_cronograma_belongs_to_obra(fields.get('cronograma_id'), obra_id)
        activity = PlanejamentoAtividade(
            obra_id=obra_id,
            orcamento_item_id=item.id,
            criado_por_user_id=user_id,
        )
        apply_activity_fields(activity, fields)
        db.session.add(activity)
        created.append(activity)
    db.session.flush()
    return created, skipped


def _normalize_header(value):
    value = unicodedata.normalize('NFKD', str(value or ''))
    value = ''.join(char for char in value if not unicodedata.combining(char))
    return re.sub(r'[^a-z0-9]+', '_', value.lower()).strip('_')


HEADER_ALIASES = {
    'titulo': {'atividade', 'titulo', 'nome', 'descricao', 'servico'},
    'descricao': {'observacao', 'observacoes', 'detalhe', 'detalhes'},
    'etapa_nome': {'etapa', 'fase', 'grupo'},
    'quantidade_planejada': {'quantidade', 'qtd', 'quantidade_planejada', 'meta'},
    'unidade': {'unidade', 'un', 'und'},
    'responsavel': {'responsavel', 'encarregado'},
    'equipe': {'equipe', 'time'},
    'data_inicio': {'inicio', 'data_inicio', 'inicia_em'},
    'data_fim': {'fim', 'termino', 'data_fim', 'termina_em'},
    'prioridade': {'prioridade'},
}


def _canonical_headers(raw_headers):
    aliases = {
        alias: canonical
        for canonical, values in HEADER_ALIASES.items()
        for alias in values
    }
    canonical = []
    for header in raw_headers:
        normalized = _normalize_header(header)
        canonical.append(aliases.get(normalized, normalized))
    if 'titulo' not in canonical:
        raise PlanejamentoValidationError(
            'A planilha precisa ter uma coluna Atividade, Nome ou Descrição.',
            'arquivo',
        )
    return canonical


def _read_csv_rows(content):
    try:
        text = content.decode('utf-8-sig')
    except UnicodeDecodeError as exc:
        raise PlanejamentoValidationError(
            'O CSV deve estar codificado em UTF-8.', 'arquivo'
        ) from exc
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        return list(csv.reader(io.StringIO(text), delimiter=';'))
    reader = csv.reader(io.StringIO(text), dialect)
    return list(reader)


def _read_xlsx_rows(content):
    if not zipfile.is_zipfile(io.BytesIO(content)):
        raise PlanejamentoValidationError('Arquivo XLSX inválido.', 'arquivo')
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        infos = archive.infolist()
        if len(infos) > 200:
            raise PlanejamentoValidationError('XLSX contém arquivos internos demais.', 'arquivo')
        uncompressed = sum(info.file_size for info in infos)
        compressed = max(1, sum(info.compress_size for info in infos))
        if (
            uncompressed > MAX_PLANILHA_UNCOMPRESSED_BYTES
            or uncompressed / compressed > 100
        ):
            raise PlanejamentoValidationError(
                'XLSX excede o limite seguro de descompactação.', 'arquivo'
            )
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            if index > MAX_PLANILHA_ROWS:
                raise PlanejamentoValidationError(
                    f'A planilha pode ter no máximo {MAX_PLANILHA_ROWS} linhas.',
                    'arquivo',
                )
            if len(row) > MAX_PLANILHA_COLUMNS and any(
                value not in (None, '') for value in row[MAX_PLANILHA_COLUMNS:]
            ):
                raise PlanejamentoValidationError(
                    f'A planilha pode ter no máximo {MAX_PLANILHA_COLUMNS} colunas.',
                    'arquivo',
                )
            rows.append(list(row[:MAX_PLANILHA_COLUMNS]))
        return rows
    finally:
        workbook.close()


def parse_spreadsheet(file_storage):
    if not file_storage or not file_storage.filename:
        raise PlanejamentoValidationError('Selecione um arquivo CSV ou XLSX.', 'arquivo')
    filename = file_storage.filename.lower().strip()
    if not filename.endswith(('.csv', '.xlsx')):
        raise PlanejamentoValidationError('Formato aceito: CSV ou XLSX.', 'arquivo')
    content = file_storage.read(MAX_PLANILHA_BYTES + 1)
    if not content:
        raise PlanejamentoValidationError('O arquivo está vazio.', 'arquivo')
    if len(content) > MAX_PLANILHA_BYTES:
        raise PlanejamentoValidationError('O arquivo deve ter no máximo 2 MB.', 'arquivo')
    rows = _read_csv_rows(content) if filename.endswith('.csv') else _read_xlsx_rows(content)
    rows = [row for row in rows if any(value not in (None, '') for value in row)]
    if len(rows) < 2:
        raise PlanejamentoValidationError('A planilha não possui atividades.', 'arquivo')
    if len(rows) > MAX_PLANILHA_ROWS + 1:
        raise PlanejamentoValidationError(
            f'A planilha pode ter no máximo {MAX_PLANILHA_ROWS} atividades.',
            'arquivo',
        )
    if any(len(row) > MAX_PLANILHA_COLUMNS for row in rows):
        raise PlanejamentoValidationError(
            f'A planilha pode ter no máximo {MAX_PLANILHA_COLUMNS} colunas.',
            'arquivo',
        )
    headers = _canonical_headers(rows[0])
    payloads = []
    errors = []
    for row_number, row in enumerate(rows[1:MAX_PLANILHA_ROWS + 1], start=2):
        raw = {
            headers[index]: value
            for index, value in enumerate(row[:len(headers)])
            if headers[index] in HEADER_ALIASES
        }
        try:
            payloads.append(normalize_activity_payload(raw, origem='planilha', partial=False))
        except PlanejamentoValidationError as exc:
            errors.append({'linha': row_number, 'campo': exc.field, 'erro': str(exc)})
    if errors:
        raise PlanejamentoValidationError(
            'A planilha contém dados inválidos.', 'arquivo', details=errors[:30]
        )
    return payloads


def serialize_preview(payloads):
    preview = []
    for payload in payloads[:20]:
        item = dict(payload)
        for field in ('data_inicio', 'data_fim'):
            if item.get(field):
                item[field] = item[field].isoformat()
        if isinstance(item.get('quantidade_planejada'), Decimal):
            item['quantidade_planejada'] = float(item['quantidade_planejada'])
        preview.append(item)
    return preview


def parse_int(value, field, minimum=None, maximum=None):
    try:
        number = int(value)
    except (TypeError, ValueError) as exc:
        raise PlanejamentoValidationError(f'{field} deve ser inteiro.', field) from exc
    if minimum is not None and number < minimum:
        raise PlanejamentoValidationError(f'{field} deve ser no mínimo {minimum}.', field)
    if maximum is not None and number > maximum:
        raise PlanejamentoValidationError(f'{field} deve ser no máximo {maximum}.', field)
    return number


def safe_percentage(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0.0
    return number if math.isfinite(number) else 0.0
