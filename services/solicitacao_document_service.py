"""Leitura e exportacao de pedidos do modulo Solicitacoes.

O leitor e deliberadamente conservador: so devolve linhas que tenham descricao
e quantidade positiva. O usuario sempre revisa os itens no formulario antes de
criar ou editar a solicitacao.
"""
from __future__ import annotations

import io
import os
import re
import unicodedata
import zipfile
from datetime import datetime
from decimal import Decimal

import pdfplumber
import reportlab
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


MAX_PEDIDO_BYTES = 10 * 1024 * 1024
MAX_PEDIDO_ITENS = 500
MAX_PEDIDO_LINHAS = 10_000
MAX_XLSX_DESCOMPACTADO = 50 * 1024 * 1024
MAX_PDF_PAGINAS = 100
FORMATOS_LEITURA = {'.xlsx', '.pdf'}

_REPORTLAB_FONTS = os.path.join(os.path.dirname(reportlab.__file__), 'fonts')
pdfmetrics.registerFont(TTFont('ObralySans', os.path.join(_REPORTLAB_FONTS, 'Vera.ttf')))
pdfmetrics.registerFont(TTFont('ObralySansBold', os.path.join(_REPORTLAB_FONTS, 'VeraBd.ttf')))
pdfmetrics.registerFontFamily('ObralySans', normal='ObralySans', bold='ObralySansBold')

_ALIASES = {
    'descricao': {
        'descricao', 'item', 'itens', 'produto', 'material', 'servico',
        'insumo', 'nome',
    },
    'quantidade': {'quantidade', 'qtd', 'qtde', 'quant', 'qte'},
    'unidade': {'unidade', 'un', 'und', 'unid', 'medida'},
    'categoria': {'categoria', 'grupo', 'familia'},
    'especificacao': {'especificacao', 'especificacoes', 'dimensao', 'dimensoes'},
    'observacao': {
        'observacao', 'obs', 'detalhe', 'detalhes', 'complemento',
        'comentario', 'comentarios',
    },
}


class PedidoLeituraError(ValueError):
    """Erro de arquivo/formato que pode ser exibido diretamente ao usuario."""


def _texto(valor):
    if valor is None:
        return ''
    return re.sub(r'\s+', ' ', str(valor).replace('\x00', ' ')).strip()


def _normalizar(valor):
    texto = unicodedata.normalize('NFKD', _texto(valor))
    texto = ''.join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r'[^a-z0-9]+', ' ', texto.lower()).strip()


def _numero(valor):
    if valor is None or isinstance(valor, bool):
        return None
    if isinstance(valor, (int, float, Decimal)):
        return float(valor)
    texto = _texto(valor).replace('R$', '').replace(' ', '')
    if not texto:
        return None
    if ',' in texto and '.' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    elif ',' in texto:
        texto = texto.replace(',', '.')
    try:
        return float(texto)
    except (TypeError, ValueError):
        return None


def _mapa_cabecalho(linha):
    mapa = {}
    for indice, valor in enumerate(linha or []):
        nome = _normalizar(valor)
        for campo, aliases in _ALIASES.items():
            if nome in aliases and campo not in mapa:
                mapa[campo] = indice
    return mapa if 'descricao' in mapa and 'quantidade' in mapa else None


def _valor_coluna(linha, indice):
    return linha[indice] if indice is not None and indice < len(linha) else None


def _montar_item(linha, mapa):
    descricao = _texto(_valor_coluna(linha, mapa.get('descricao')))
    quantidade = _numero(_valor_coluna(linha, mapa.get('quantidade')))
    if not descricao or quantidade is None or quantidade <= 0:
        return None
    if _normalizar(descricao) in {'total', 'subtotal', 'valor total'}:
        return None
    unidade = _texto(_valor_coluna(linha, mapa.get('unidade')))
    observacao = _texto(_valor_coluna(linha, mapa.get('observacao')))
    categoria = _texto(_valor_coluna(linha, mapa.get('categoria')))
    especificacao = _texto(_valor_coluna(linha, mapa.get('especificacao')))
    complementos = []
    if categoria:
        complementos.append(f'Categoria: {categoria}')
    if especificacao:
        complementos.append(f'Especificação: {especificacao}')
    if observacao:
        complementos.append(observacao)
    observacao = ' | '.join(complementos)
    return {
        'descricao': descricao[:300],
        'quantidade': quantidade,
        'unidade': unidade[:20],
        'observacao': observacao[:300],
    }


def _itens_de_linhas(linhas, permitir_sem_cabecalho=True):
    linhas = [list(linha or []) for linha in linhas]
    inicio = None
    mapa = None
    for indice, linha in enumerate(linhas[:30]):
        encontrado = _mapa_cabecalho(linha)
        if encontrado:
            inicio, mapa = indice + 1, encontrado
            break

    if mapa is None and permitir_sem_cabecalho:
        # Formato simples sem titulo: descricao | quantidade | unidade | obs.
        mapa = {'descricao': 0, 'quantidade': 1, 'unidade': 2, 'observacao': 3}
        inicio = 0
    if mapa is None:
        return []

    itens = []
    for linha in linhas[inicio:]:
        if _mapa_cabecalho(linha):  # cabecalho repetido entre paginas
            continue
        item = _montar_item(linha, mapa)
        if item:
            itens.append(item)
        if len(itens) >= MAX_PEDIDO_ITENS:
            break
    return itens


def _ler_xlsx(conteudo):
    if not conteudo.startswith(b'PK'):
        raise PedidoLeituraError('O arquivo não é um Excel .xlsx válido.')
    try:
        with zipfile.ZipFile(io.BytesIO(conteudo)) as pacote:
            arquivos = pacote.infolist()
            tamanho_total = sum(item.file_size for item in arquivos)
            if len(arquivos) > 1_000 or tamanho_total > MAX_XLSX_DESCOMPACTADO:
                raise PedidoLeituraError('A planilha é grande ou complexa demais para leitura segura.')
    except PedidoLeituraError:
        raise
    except (OSError, zipfile.BadZipFile) as exc:
        raise PedidoLeituraError('O arquivo não é um Excel .xlsx válido.') from exc
    try:
        workbook = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as exc:
        raise PedidoLeituraError('Não foi possível abrir a planilha. Salve-a como .xlsx e tente novamente.') from exc

    avisos = []
    for planilha in workbook.worksheets:
        linhas = list(planilha.iter_rows(
            min_row=1,
            max_row=min(planilha.max_row or MAX_PEDIDO_LINHAS, MAX_PEDIDO_LINHAS),
            max_col=min(planilha.max_column or 50, 50),
            values_only=True,
        ))
        itens = _itens_de_linhas(linhas)
        if itens:
            if (planilha.max_row or 0) > MAX_PEDIDO_LINHAS:
                avisos.append(f'Foram analisadas somente as primeiras {MAX_PEDIDO_LINHAS} linhas da planilha.')
            if len(itens) == MAX_PEDIDO_ITENS:
                avisos.append(f'Foram importados somente os primeiros {MAX_PEDIDO_ITENS} itens.')
            return itens, avisos, planilha.title
    raise PedidoLeituraError(
        'Nenhum item foi identificado. Use colunas “Descrição” e “Quantidade”; “Unidade” e “Observação” são opcionais.'
    )


def _ler_pdf(conteudo):
    if not conteudo.lstrip().startswith(b'%PDF'):
        raise PedidoLeituraError('O arquivo não é um PDF válido.')
    avisos = []
    try:
        with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
            if len(pdf.pages) > MAX_PDF_PAGINAS:
                raise PedidoLeituraError(f'O PDF ultrapassa o limite de {MAX_PDF_PAGINAS} páginas.')
            itens = []
            for pagina in pdf.pages:
                for tabela in pagina.extract_tables() or []:
                    encontrados = _itens_de_linhas(tabela, permitir_sem_cabecalho=False)
                    itens.extend(encontrados)
                    if len(itens) >= MAX_PEDIDO_ITENS:
                        break
                if len(itens) >= MAX_PEDIDO_ITENS:
                    break

            if not itens:
                # Fallback para PDFs gerados sem bordas de tabela. Mantemos os
                # espacos do layout para reconstruir as quatro colunas.
                linhas = []
                for pagina in pdf.pages:
                    texto = pagina.extract_text(layout=True) or ''
                    for linha in texto.splitlines():
                        colunas = [c.strip() for c in re.split(r'\s{2,}', linha.strip()) if c.strip()]
                        if len(colunas) >= 2:
                            linhas.append(colunas)
                itens = _itens_de_linhas(linhas, permitir_sem_cabecalho=False)
    except PedidoLeituraError:
        raise
    except Exception as exc:
        raise PedidoLeituraError('Não foi possível ler o PDF enviado.') from exc

    if not itens:
        raise PedidoLeituraError(
            'Nenhum item foi identificado no PDF. Envie um PDF pesquisável com tabela e colunas “Descrição” e “Quantidade”, ou use Excel.'
        )
    if len(itens) >= MAX_PEDIDO_ITENS:
        itens = itens[:MAX_PEDIDO_ITENS]
        avisos.append(f'Foram importados somente os primeiros {MAX_PEDIDO_ITENS} itens.')
    return itens, avisos, None


def ler_pedido(arquivo):
    nome = _texto(getattr(arquivo, 'filename', ''))
    if not nome:
        raise PedidoLeituraError('Selecione um arquivo Excel (.xlsx) ou PDF.')
    extensao = '.' + nome.rsplit('.', 1)[-1].lower() if '.' in nome else ''
    if extensao not in FORMATOS_LEITURA:
        raise PedidoLeituraError('Formato não aceito. Envie Excel (.xlsx) ou PDF.')

    conteudo = arquivo.stream.read(MAX_PEDIDO_BYTES + 1)
    if not conteudo:
        raise PedidoLeituraError('O arquivo enviado está vazio.')
    if len(conteudo) > MAX_PEDIDO_BYTES:
        raise PedidoLeituraError('O arquivo ultrapassa o limite de 10 MB.')

    if extensao == '.xlsx':
        itens, avisos, planilha = _ler_xlsx(conteudo)
    else:
        itens, avisos, planilha = _ler_pdf(conteudo)
    return {
        'arquivo': nome,
        'formato': extensao[1:],
        'planilha': planilha,
        'quantidade_itens': len(itens),
        'itens': itens,
        'avisos': avisos,
    }


def _data_br(valor, com_hora=False):
    if not valor:
        return '—'
    if isinstance(valor, str):
        try:
            valor = datetime.fromisoformat(valor.replace('Z', '+00:00'))
        except ValueError:
            return valor
    formato = '%d/%m/%Y %H:%M' if com_hora else '%d/%m/%Y'
    return valor.strftime(formato)


def _quantidade_excel(valor):
    numero = float(valor or 0)
    return int(numero) if numero.is_integer() else numero


def gerar_xlsx(solicitacao):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedido'
    azul = '0F766E'
    azul_escuro = '0F172A'
    cinza = 'E2E8F0'
    borda = Side(style='thin', color=cinza)

    ws.merge_cells('A1:E1')
    ws['A1'] = f'PEDIDO DE COTAÇÃO — SOLICITAÇÃO #{solicitacao.id}'
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=azul_escuro)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    metadados = [
        ('Obra', solicitacao.obra.nome if solicitacao.obra else '—'),
        ('Solicitante', solicitacao.solicitante.username if solicitacao.solicitante else '—'),
        ('Solicitada em', _data_br(solicitacao.data_criacao, com_hora=True)),
        ('Necessidade', _data_br(solicitacao.data_necessidade)),
        ('Tipo', solicitacao.tipo or '—'),
        ('Observação geral', solicitacao.observacao or '—'),
    ]
    linha = 3
    for rotulo, valor in metadados:
        ws[f'A{linha}'] = rotulo
        ws[f'A{linha}'].font = Font(bold=True, color=azul_escuro)
        ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=5)
        ws[f'B{linha}'] = valor
        ws[f'B{linha}'].alignment = Alignment(wrap_text=True)
        linha += 1

    linha += 1
    cabecalho = ['Item', 'Descrição', 'Quantidade', 'Unidade', 'Observação']
    for coluna, titulo in enumerate(cabecalho, start=1):
        celula = ws.cell(linha, coluna, titulo)
        celula.font = Font(bold=True, color='FFFFFF')
        celula.fill = PatternFill('solid', fgColor=azul)
        celula.alignment = Alignment(horizontal='center')
        celula.border = Border(bottom=borda)
    linha_cabecalho = linha

    for indice, item in enumerate(solicitacao.itens, start=1):
        linha += 1
        valores = [
            indice, item.descricao, _quantidade_excel(item.quantidade),
            item.unidade or '', item.observacao or '',
        ]
        for coluna, valor in enumerate(valores, start=1):
            celula = ws.cell(linha, coluna, valor)
            celula.border = Border(bottom=borda)
            celula.alignment = Alignment(
                horizontal='center' if coluna in (1, 3, 4) else 'left',
                vertical='top', wrap_text=True,
            )
        ws.cell(linha, 3).number_format = '#,##0.00'

    larguras = [16, 42, 14, 14, 42]
    for indice, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = largura
    ws.freeze_panes = f'A{linha_cabecalho + 1}'
    ws.auto_filter.ref = f'A{linha_cabecalho}:E{linha}'
    ws.print_title_rows = f'{linha_cabecalho}:{linha_cabecalho}'
    ws.print_area = f'A1:E{linha}'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    saida = io.BytesIO()
    wb.save(saida)
    return saida.getvalue()


def _texto_pdf(valor):
    return _texto(valor).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;') or '—'


def gerar_pdf(solicitacao):
    saida = io.BytesIO()
    doc = SimpleDocTemplate(
        saida, pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title=f'Solicitação #{solicitacao.id}', author='Obraly',
    )
    estilos = getSampleStyleSheet()
    titulo = ParagraphStyle(
        'TituloPedido', parent=estilos['Heading1'], fontName='ObralySansBold',
        fontSize=16, leading=20, textColor=colors.HexColor('#0F172A'),
        alignment=TA_CENTER, spaceAfter=10,
    )
    pequeno = ParagraphStyle(
        'PequenoPedido', parent=estilos['BodyText'], fontName='ObralySans',
        fontSize=8.5, leading=11, textColor=colors.HexColor('#334155'),
        alignment=TA_LEFT,
    )
    cab = ParagraphStyle(
        'CabPedido', parent=pequeno, fontName='ObralySansBold',
        textColor=colors.white, alignment=TA_CENTER,
    )
    corpo = ParagraphStyle('CorpoPedido', parent=pequeno, fontSize=8, leading=10)

    historia = [Paragraph(f'PEDIDO DE COTAÇÃO — SOLICITAÇÃO #{solicitacao.id}', titulo)]
    meta = [
        [Paragraph('<b>Obra</b>', pequeno), Paragraph(_texto_pdf(solicitacao.obra.nome if solicitacao.obra else None), pequeno),
         Paragraph('<b>Tipo</b>', pequeno), Paragraph(_texto_pdf(solicitacao.tipo), pequeno)],
        [Paragraph('<b>Solicitante</b>', pequeno), Paragraph(_texto_pdf(solicitacao.solicitante.username if solicitacao.solicitante else None), pequeno),
         Paragraph('<b>Solicitada em</b>', pequeno), Paragraph(_data_br(solicitacao.data_criacao, True), pequeno)],
        [Paragraph('<b>Necessidade</b>', pequeno), Paragraph(_data_br(solicitacao.data_necessidade), pequeno),
         Paragraph('<b>Observação geral</b>', pequeno), Paragraph(_texto_pdf(solicitacao.observacao), pequeno)],
    ]
    tabela_meta = Table(meta, colWidths=[25 * mm, 58 * mm, 30 * mm, 65 * mm])
    tabela_meta.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (-1, -1), .5, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0, 0), (-1, -1), .25, colors.HexColor('#E2E8F0')),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    historia.extend([tabela_meta, Spacer(1, 8 * mm)])

    dados = [[
        Paragraph('Item', cab), Paragraph('Descrição', cab),
        Paragraph('Quantidade', cab), Paragraph('Unidade', cab),
        Paragraph('Observação', cab),
    ]]
    for indice, item in enumerate(solicitacao.itens, start=1):
        quantidade = _quantidade_excel(item.quantidade)
        quantidade_txt = str(quantidade).replace('.', ',')
        dados.append([
            Paragraph(str(indice), corpo), Paragraph(_texto_pdf(item.descricao), corpo),
            Paragraph(quantidade_txt, corpo), Paragraph(_texto_pdf(item.unidade), corpo),
            Paragraph(_texto_pdf(item.observacao), corpo),
        ])
    tabela = Table(dados, colWidths=[11 * mm, 65 * mm, 25 * mm, 23 * mm, 54 * mm], repeatRows=1)
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),
        ('BOX', (0, 0), (-1, -1), .5, colors.HexColor('#94A3B8')),
        ('INNERGRID', (0, 0), (-1, -1), .25, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    historia.extend([
        tabela,
        Spacer(1, 6 * mm),
        Paragraph('Documento gerado pelo Obraly para cotação com fornecedores.', pequeno),
    ])
    doc.build(historia)
    return saida.getvalue()
